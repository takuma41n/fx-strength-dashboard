"""データ取得層.

すべて無料・APIキー不要のソース。各フェッチャは
    {"YYYY-MM-DD": float, ...}
という形の系列を返す。取得に失敗した場合は FetchError を送出し、
呼び出し側（fetch_market.py）が前回値へフォールバックする。
"""

from __future__ import annotations

import csv
import datetime as dt
import io
import json
import re
import time
import urllib.parse
import urllib.request
import zipfile

UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)


class FetchError(Exception):
    pass


def http_get(url, *, timeout=90, retries=3, backoff=2.5, headers=None, browser=True):
    """リトライ付きの GET。

    ホストによって求められるものが正反対なので browser フラグで切り替える:
      * RBA / BOE / 財務省 — ブラウザ相当のヘッダがないと弾かれる
      * FRED — 逆にブラウザ風 User-Agent を付けると応答が返ってこなくなる
        （fredgraph.csv はヘッダ無しの素の urllib なら即座に返る）
    """
    hdrs = {}
    if browser:
        hdrs = {
            "User-Agent": UA,
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9,ja;q=0.8",
            "Connection": "close",
        }
    if headers:
        hdrs.update(headers)
    last = None
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers=hdrs)
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                return resp.read()
        except Exception as exc:  # noqa: BLE001 - あらゆる失敗をリトライ対象にする
            last = exc
            if attempt < retries - 1:
                time.sleep(backoff * (attempt + 1))
    raise FetchError(f"{url} -> {last}")


def _f(value):
    """数値化。欠損表現（'', '.', '-', 'NA'）は None。"""
    if value is None:
        return None
    s = str(value).strip().replace(",", "")
    if s in ("", ".", "-", "NA", "na", "null", "None"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _looks_like_html(raw: bytes) -> bool:
    head = raw[:400].lstrip().lower()
    return head.startswith(b"<!doctype") or head.startswith(b"<html") or b"<script" in head


# ---------------------------------------------------------------------------
# Yahoo Finance — FX・株価指数・商品・VIX・FF先物
# ---------------------------------------------------------------------------

def fetch_yahoo(ticker: str, range_: str = "2y") -> dict:
    url = (
        f"https://query1.finance.yahoo.com/v8/finance/chart/{urllib.parse.quote(ticker)}"
        f"?range={range_}&interval=1d"
    )
    raw = http_get(url, timeout=45)
    try:
        payload = json.loads(raw)
        result = payload["chart"]["result"][0]
    except Exception as exc:  # noqa: BLE001
        raise FetchError(f"Yahoo {ticker}: レスポンスを解釈できない ({exc})") from exc

    stamps = result.get("timestamp") or []
    closes = result["indicators"]["quote"][0].get("close") or []
    out = {}
    for ts, close in zip(stamps, closes):
        val = _f(close)
        if val is None:
            continue
        date = dt.datetime.utcfromtimestamp(ts).date().isoformat()
        out[date] = val
    if not out:
        raise FetchError(f"Yahoo {ticker}: 有効な終値が0件")
    return out


# ---------------------------------------------------------------------------
# WTI 2番限 — Yahoo は限月コード（CLV26.NYM 等）を個別に配信している
# ---------------------------------------------------------------------------

# 先物の限月コード（1月=F 〜 12月=Z）
FUTURES_MONTH_CODES = "FGHJKMNQUVXZ"


def _cl_ticker(year: int, month: int) -> str:
    return f"CL{FUTURES_MONTH_CODES[month - 1]}{year % 100:02d}.NYM"


def fetch_wti_second_month(recent: int = 5) -> dict:
    """期近(CL=F)の次の限月の終値を返す。

    Yahoo は CL=F を「その時の期近」として配信しており、実測で CL=F の終値は
    期近限月（例: CLU26.NYM）と完全一致する。そこで翌月以降の限月を順に叩き、
    CL=F と一致したものを期近と同定し、その次に取れた限月を2番限とする。
    限月コードから機械的に決めると限月交代日（受渡月前月25日の3営業日前）の
    祝日判定が必要になるが、同定方式なら交代日を知らなくても自己修復する。

    末尾 recent 日分しか返さないのが肝。fetch_market.py の merge_series は
    dict.update なので、全履歴を返すと限月交代のたびに過去の2番限価格が
    新しい限月の履歴で上書きされ、ロール系列（継ぎ接ぎの2番限系列）が壊れる。
    """
    front = fetch_yahoo("CL=F", range_="1mo")
    front_last_date = max(front)
    front_last = front[front_last_date]
    today = dt.date.today()

    passed_front = False
    for k in range(1, 8):
        year = today.year + (today.month - 1 + k) // 12
        month = (today.month - 1 + k) % 12 + 1
        try:
            cand = fetch_yahoo(_cl_ticker(year, month), range_="1mo")
        except FetchError:
            continue
        finally:
            time.sleep(1.0)  # Yahoo のレート制限に余裕を持たせる
        cand_last_date = max(cand)
        # 配信が5日以上前で止まっている限月は満期切れとみなしてスキップ
        gap = (dt.date.fromisoformat(front_last_date)
               - dt.date.fromisoformat(cand_last_date)).days
        if gap > 5:
            continue
        if not passed_front:
            if abs(cand[cand_last_date] - front_last) <= 0.005:
                passed_front = True
            continue
        # 期近の次に取れた有効な限月 = 2番限
        keys = sorted(cand)[-recent:]
        return {d: cand[d] for d in keys}
    raise FetchError("WTI 2番限: CL=F と一致する期近限月を同定できない")


def fetch_wti_m2_history(months_back: int = 20) -> dict:
    """過去の各限月から「2番限だった期間」だけを切り出して縫い合わせる。

    bootstrap_history.py からのみ呼ぶ（一度きり）。受渡月 D の限月が
    2番限なのはおおむね「(D-3)月の20日 〜 (D-2)月の20日」
    （限月交代は受渡月前月25日の3営業日前ごろ）。境界の1〜2日のずれは
    z-score に実質影響しない。満期済み限月を Yahoo が配信していない場合は
    その月だけ欠けるが、z の計算はできる。
    """
    def _day20(year: int, month: int, delta_months: int) -> str:
        y = year + (month - 1 + delta_months) // 12
        m = (month - 1 + delta_months) % 12 + 1
        return dt.date(y, m, 20).isoformat()

    today = dt.date.today()
    out = {}
    for k in range(-months_back, 3):
        year = today.year + (today.month - 1 + k) // 12
        month = (today.month - 1 + k) % 12 + 1
        lo = _day20(year, month, -3)
        hi = _day20(year, month, -2)
        try:
            cand = fetch_yahoo(_cl_ticker(year, month), range_="2y")
        except FetchError:
            continue
        finally:
            time.sleep(1.0)
        for d, v in cand.items():
            if lo <= d < hi:
                out[d] = v
    if not out:
        raise FetchError("WTI 2番限履歴: 有効値が0件")
    return out


# ---------------------------------------------------------------------------
# FRED — APIキー不要の fredgraph.csv を使う
# ---------------------------------------------------------------------------

def fetch_fred(series_id: str) -> dict:
    raw = http_get(f"https://fred.stlouisfed.org/graph/fredgraph.csv?id={series_id}",
                   timeout=60, browser=False)
    if _looks_like_html(raw):
        raise FetchError(f"FRED {series_id}: HTMLが返った（系列名が無効の可能性）")
    reader = csv.reader(io.StringIO(raw.decode("utf-8-sig", errors="replace")))
    rows = [r for r in reader if r]
    if len(rows) < 2:
        raise FetchError(f"FRED {series_id}: データ行なし")
    out = {}
    for row in rows[1:]:
        if len(row) < 2:
            continue
        val = _f(row[1])
        if val is None:
            continue
        out[row[0].strip()] = val
    if not out:
        raise FetchError(f"FRED {series_id}: 有効値が0件")
    return out


# ---------------------------------------------------------------------------
# ECB Data Portal — ユーロ圏2年AAAスポットレート（日次）
# ---------------------------------------------------------------------------

ECB_2Y_KEY = "YC/B.U2.EUR.4F.G_N_A.SV_C_YM.SR_2Y"


def fetch_ecb_2y(observations: int = 600) -> dict:
    url = (
        f"https://data-api.ecb.europa.eu/service/data/{ECB_2Y_KEY}"
        f"?lastNObservations={observations}&format=csvdata&detail=dataonly"
    )
    raw = http_get(url, timeout=60)
    reader = csv.DictReader(io.StringIO(raw.decode("utf-8-sig", errors="replace")))
    out = {}
    for row in reader:
        date = (row.get("TIME_PERIOD") or "").strip()
        val = _f(row.get("OBS_VALUE"))
        if date and val is not None:
            out[date] = val
    if not out:
        raise FetchError("ECB 2Y: 有効値が0件")
    return out


# ---------------------------------------------------------------------------
# 財務省 — 日本国債金利（Shift-JIS・和暦）
#   data/jgbcm_all.csv : 1974年〜前月末までの全履歴
#   jgbcm.csv          : 当月分
# ---------------------------------------------------------------------------

MOF_BASE = "https://www.mof.go.jp/jgbs/reference/interest_rate/"
_ERA_BASE = {"M": 1867, "T": 1911, "S": 1925, "H": 1988, "R": 2018}
_WAREKI = re.compile(r"^([MTSHR])(\d+)\.(\d+)\.(\d+)$")


def _parse_wareki(text: str):
    m = _WAREKI.match((text or "").strip())
    if not m:
        return None
    era, year, month, day = m.group(1), int(m.group(2)), int(m.group(3)), int(m.group(4))
    try:
        return dt.date(_ERA_BASE[era] + year, month, day).isoformat()
    except ValueError:
        return None


def _parse_mof_csv(raw: bytes, tenor: str = "2年") -> dict:
    text = raw.decode("shift_jis", errors="replace")
    rows = list(csv.reader(io.StringIO(text)))
    header_idx = None
    for i, row in enumerate(rows[:10]):
        if row and row[0].strip() == "基準日":
            header_idx = i
            break
    if header_idx is None:
        raise FetchError("MOF: ヘッダ行『基準日』が見つからない")
    header = [c.strip() for c in rows[header_idx]]
    if tenor not in header:
        raise FetchError(f"MOF: 列『{tenor}』が見つからない (header={header[:6]})")
    col = header.index(tenor)

    out = {}
    for row in rows[header_idx + 1:]:
        if len(row) <= col:
            continue
        date = _parse_wareki(row[0])
        if date is None:
            continue
        val = _f(row[col])
        if val is not None:
            out[date] = val
    return out


def fetch_mof_jgb(tenor: str = "2年", include_history: bool = True) -> dict:
    out = {}
    if include_history:
        try:
            out.update(_parse_mof_csv(http_get(MOF_BASE + "data/jgbcm_all.csv", timeout=120), tenor))
        except FetchError:
            pass  # 当月分だけでも取れれば既存の蓄積とマージできる
    out.update(_parse_mof_csv(http_get(MOF_BASE + "jgbcm.csv", timeout=60), tenor))
    if not out:
        raise FetchError("MOF JGB: 有効値が0件")
    return out


# ---------------------------------------------------------------------------
# EIA — 米商業原油在庫（週次、千バレル）
#   APIキーが要るのは api.eia.gov v2 のほう。データブラウザの LeafHandler は
#   キーレスで全履歴のHTMLを返すので、そちらを正規表現で読む。
#   .xls 版（hist_xls / ir.eia.gov）は旧BIFF形式で openpyxl では開けない。
# ---------------------------------------------------------------------------

EIA_LEAF = ("https://www.eia.gov/dnav/pet/hist/LeafHandler.ashx"
            "?n=PET&s=WCESTUS1&f=W")
_EIA_MONTHS = {m: i + 1 for i, m in enumerate(
    ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
     "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"])}
_EIA_ROW_HEAD = re.compile(r"(\d{4})-([A-Z][a-z]{2})")
_EIA_OBS = re.compile(r"(\d{2})/(\d{2})\s+([\d,]+(?:\.\d+)?)")


def fetch_eia_crude_stocks() -> dict:
    """行が「1982-Aug  08/20 338,764  08/27 336,138 …」の形のHTML表を読む。

    観測日は月/日しか持たないため年は行ラベルから取るが、週の区切りが
    月を跨ぐと隣の月（年末は隣の年）の日付が同じ行に載る。月の差が
    6を超えたら年を±1して補正する。
    """
    raw = http_get(EIA_LEAF, timeout=90)
    text = raw.decode("utf-8", errors="replace")
    out = {}
    for tr in re.findall(r"<tr[^>]*>(.*?)</tr>", text, flags=re.S | re.I):
        plain = re.sub(r"<[^>]+>", " ", tr).replace("&nbsp;", " ")
        plain = re.sub(r"\s+", " ", plain).strip()
        head = _EIA_ROW_HEAD.match(plain)
        if not head:
            continue
        row_year = int(head.group(1))
        row_month = _EIA_MONTHS.get(head.group(2))
        if row_month is None:
            continue
        for mm, dd, val in _EIA_OBS.findall(plain, head.end()):
            month, day = int(mm), int(dd)
            year = row_year
            if month - row_month > 6:
                year -= 1
            elif row_month - month > 6:
                year += 1
            value = _f(val)
            if value is None:
                continue
            try:
                out[dt.date(year, month, day).isoformat()] = value
            except ValueError:
                continue
    if not out:
        raise FetchError("EIA WCESTUS1: 有効値が0件（HTML構造が変わった可能性）")
    return out


# ---------------------------------------------------------------------------
# BOE — 英国スポットカーブ / OISカーブ（日次）
#   latest-yield-curve-data.zip : 当月分（約400KB、日次CI向け）
#   glcnominalddata.zip         : 全履歴（約39MB、初回ブートストラップのみ）
#   oisddata.zip                : OIS全履歴
# ---------------------------------------------------------------------------

BOE_BASE = "https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/"

# Excel のシリアル日付起点（1900年方式、Lotus互換のうるう年バグ込み）
_EXCEL_EPOCH = dt.date(1899, 12, 30)


def _excel_date(value):
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, (int, float)) and value > 20000:
        return (_EXCEL_EPOCH + dt.timedelta(days=int(value))).isoformat()
    return None


def _extract_curve_from_workbook(data: bytes, tenor: float = 2.0) -> dict:
    """BOEカーブの xlsx から指定年限の系列を取り出す。

    シート構成が変わっても壊れないよう、シート名に spot/fwd を仮定せず
    『数値の年限が並んだヘッダ行』を総当たりで探す。
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    candidates = [n for n in wb.sheetnames if "spot" in n.lower()] or list(wb.sheetnames)

    for name in candidates:
        ws = wb[name]
        rows = ws.iter_rows(values_only=True)
        header, col = None, None
        for row in rows:
            if row is None:
                continue
            numeric = {}
            for idx, cell in enumerate(row):
                if isinstance(cell, (int, float)) and 0 < cell <= 60:
                    numeric[idx] = float(cell)
            # ヘッダ行は「年限のはしご」なので単調増加する。これを条件に入れないと
            # 利回りが並んだデータ行をヘッダと誤認する。
            if len(numeric) >= 5:
                vals = [numeric[i] for i in sorted(numeric)]
                if all(b > a for a, b in zip(vals, vals[1:])):
                    match = [i for i, v in numeric.items() if abs(v - tenor) < 1e-6]
                    if match:
                        header, col = numeric, match[0]
                        break
        if header is None or col is None:
            continue

        out = {}
        for row in rows:  # ヘッダ行の続きから読む
            if row is None or len(row) <= col:
                continue
            date = _excel_date(row[0])
            if date is None:
                continue
            val = _f(row[col])
            if val is not None:
                out[date] = val
        if out:
            wb.close()
            return out
    wb.close()
    raise FetchError(f"BOE: 年限{tenor}の列を含むシートが見つからない")


def _boe_from_zip(zip_bytes: bytes, name_filter: str, tenor: float) -> dict:
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    names = [n for n in zf.namelist() if n.lower().endswith((".xlsx", ".xls"))]
    targets = [n for n in names if name_filter.lower() in n.lower()] or names
    merged = {}
    errors = []
    for name in targets:
        try:
            merged.update(_extract_curve_from_workbook(zf.read(name), tenor))
        except Exception as exc:  # noqa: BLE001
            errors.append(f"{name}: {exc}")
    if not merged:
        raise FetchError(f"BOE zip: 抽出失敗 ({'; '.join(errors) or 'ファイルなし'})")
    return merged


def fetch_boe_2y_current() -> dict:
    """当月分のみ（日次CI用・軽量）。履歴は data.json の蓄積とマージする。"""
    raw = http_get(BOE_BASE + "latest-yield-curve-data.zip", timeout=120)
    return _boe_from_zip(raw, "nominal", 2.0)


def fetch_boe_2y_history() -> dict:
    """全履歴（約39MB）。bootstrap_history.py からのみ呼ぶ。"""
    raw = http_get(BOE_BASE + "glcnominalddata.zip", timeout=600)
    return _boe_from_zip(raw, "", 2.0)


def fetch_boe_ois_1y_current() -> dict:
    """英OISカーブ1年（政策織り込みの参考表示用）。"""
    raw = http_get(BOE_BASE + "latest-yield-curve-data.zip", timeout=120)
    return _boe_from_zip(raw, "ois", 1.0)


# ---------------------------------------------------------------------------
# RBA — 豪州2年国債（bot判定が厳しく、ローカルIPからは弾かれることがある）
# ---------------------------------------------------------------------------

# f2.1 は同じ列構成でも月次。日次が必要なので f2 を使う。
RBA_F2 = "https://www.rba.gov.au/statistics/tables/csv/f2-data.csv"
_RBA_DATE_FORMATS = ("%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%b-%y")


def _parse_rba_date(text: str):
    s = (text or "").strip()
    for fmt in _RBA_DATE_FORMATS:
        try:
            return dt.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def fetch_rba_2y() -> dict:
    # ブラウザ相当の User-Agent を付けると 403 で弾かれる。素の urllib なら通る。
    raw = http_get(RBA_F2, timeout=90, retries=4, backoff=4.0, browser=False)
    if _looks_like_html(raw):
        raise FetchError("RBA: HTMLが返った（bot判定）")
    rows = list(csv.reader(io.StringIO(raw.decode("utf-8-sig", errors="replace"))))
    title_row = None
    for row in rows[:15]:
        if row and row[0].strip().lower() == "title":
            title_row = row
            break
    col = 1
    if title_row:
        for idx, cell in enumerate(title_row):
            if "2 year" in (cell or "").lower():
                col = idx
                break
    out = {}
    for row in rows:
        if len(row) <= col:
            continue
        date = _parse_rba_date(row[0])
        if date is None:
            continue
        val = _f(row[col])
        if val is not None:
            out[date] = val
    if not out:
        raise FetchError("RBA: 有効値が0件")
    return out


# ---------------------------------------------------------------------------
# SNB — SARON（CHFは2年国債の無料日次ソースが全滅したため代替）
#   zirepo キューブ: H0=SARON O/N, H7=3ヶ月コンパウンド, H8=6ヶ月コンパウンド
# ---------------------------------------------------------------------------

SNB_ZIREPO = "https://data.snb.ch/api/cube/zirepo/data/csv/en"


def fetch_snb_saron(dimension: str = "H8") -> dict:
    raw = http_get(SNB_ZIREPO, timeout=90)
    text = raw.decode("utf-8-sig", errors="replace")
    if text.lstrip().startswith("{"):
        raise FetchError(f"SNB zirepo: {text[:120]}")
    out = {}
    for line in text.splitlines():
        parts = [p.strip().strip('"') for p in line.split(";")]
        if len(parts) < 3 or parts[1] != dimension:
            continue
        date, val = parts[0], _f(parts[2])
        if val is not None and re.match(r"^\d{4}-\d{2}-\d{2}$", date):
            out[date] = val
    if not out:
        raise FetchError(f"SNB zirepo: 次元{dimension}の有効値が0件")
    return out


# ---------------------------------------------------------------------------
# 系列ユーティリティ
# ---------------------------------------------------------------------------

def merge_series(old: dict, new: dict) -> dict:
    """既存の蓄積に新規取得をかぶせる。BOE/MOF のように当月分しか
    返さないソースでも、data.json 側に履歴が溜まっていく。"""
    merged = dict(old or {})
    merged.update(new or {})
    return merged


def trim_series(series: dict, keep: int) -> dict:
    if not series:
        return {}
    keys = sorted(series)[-keep:]
    return {k: series[k] for k in keys}


def last_date(series: dict):
    return max(series) if series else None


def stale_days(series: dict, today: dt.date | None = None):
    latest = last_date(series)
    if latest is None:
        return None
    today = today or dt.date.today()
    return (today - dt.date.fromisoformat(latest)).days
